import pandas as pd
import numpy as np
import jenkspy
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================
PATH = Path(".")
CUSTOMER_FILE, RM_MASTER_FILE = PATH / "customers.xlsx", PATH / "rm_master.xlsx"
T_DATA_FILE, BRANCH_LOOKUP_FILE = PATH / "t_data.xlsx", PATH / "branch_lookup.xlsx"
OUTPUT_FILE = PATH / "assignment_recommendations.xlsx"

VACANCY_CODES = {"00101":"BM1","00102":"BS1","00103":"SSB1","00104":"SSB2",
                 "00105":"SSA1","00106":"SSA2","00107":"SSA3"}
CONTROL_CODE, EMPLOYEE_CODE = "00001", "14100"
STAFFED_ROLES = ["BM","BS","SSB","SSA","BM1","BS1","SSB1","SSB2","SSA1","SSA2","SSA3"]
ALL_SPECIAL_CODES = [CONTROL_CODE, EMPLOYEE_CODE]
HNW_THRESHOLD = 500000
CAP = {"Branch Manager":150, "Branch Supervisor":150, "SSB":200, "SSA":125}
ROLE_CASCADE = {
    "BM > BS > SSB": ("Branch Manager", ["Branch Supervisor","SSB"]),
    "BM > BS": ("Branch Manager", ["Branch Supervisor"]),
    "SSB": ("SSB",[]), "SSA": ("SSA",[]),
}
EXPECTED_TIERS = {
    "BM > BS > SSB": ["Branch Manager","Branch Supervisor","SSB"],
    "BM > BS": ["Branch Manager","Branch Supervisor"],
    "SSB": ["SSB"], "SSA": ["SSA"],
}

C = {
    "ck":"Cust Key","cn":"Cust Name","cf":"Cust Flag",
    "cc":"Cust Class Description","ef":"Cust Employee Flag",
    "bc":"Cust Branch Code","po":"Cust Primary Officer Code",
    "pd":"Cust Primary Officer Description","qb":"Cust Quarterback Officer Code",
    "cs":"Cust Cross-Sell Ratio",
    "bd":"Cust Ledger Balance Deposits","bl":"Cust Ledger Balance Loans",
    "be":"Cust Ledger Balance Equity","bi":"Cust Ledger Balance Insurance",
    "rk":"Relationship Key","rn":"Relationship Name",
    "oc":"PrimaryOfficerCode","on":"PrimaryOfficer Name",
    "rc":"RoleCode","cc2":"CostCenterCode","rb":"BranchCode","tf":"Terminated Flag",
    "tb":"Cust Primary Servicing Branch Code","tbf":"Cust Primary Branch Code",
    "bbc":"Branch Code","bcc":"Cost Center","bso":"Secondary Officer",
}

# =============================================================================
# STYLING
# =============================================================================
HF = Font(bold=True, size=11, name="Arial", color="FFFFFF")
DF = Font(size=11, name="Arial")
BD = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
FILLS = {"blue":PatternFill("solid",fgColor="4472C4"),"red":PatternFill("solid",fgColor="C00000"),
         "green":PatternFill("solid",fgColor="548235"),"orange":PatternFill("solid",fgColor="ED7D31")}

# =============================================================================
# HELPERS — clean_str only, no data alteration
# =============================================================================
def clean_num(s):
    c = s.fillna("").astype(str).str.strip().replace({"-":"0","":"0"})
    return pd.to_numeric(c.str.replace(",","",regex=False).str.replace("$","",regex=False),errors="coerce").fillna(0)

def clean_str(s):
    """Strip whitespace and hidden characters only. No numeric conversion, no leading zero removal."""
    return s.fillna("").astype(str).str.strip().str.replace("\u200b","",regex=False).str.replace("\xa0","",regex=False)

def role_tier(r):
    r = str(r).strip().upper()
    if r in ["BM","BM1"]: return "Branch Manager"
    if r in ["BS","BS1"]: return "Branch Supervisor"
    if r in ["SSB","SSB1","SSB2"]: return "SSB"
    if r in ["SSA","SSA1","SSA2","SSA3"]: return "SSA"
    return "Other"

def write_sheet(ws, headers, rows, fill=None):
    fill = fill or FILLS["blue"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.border, c.alignment = HF, fill, BD, Alignment(horizontal="center",vertical="center",wrap_text=True)
    for ri, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font, c.border = DF, BD
    for ci in range(1, len(headers)+1):
        mx = max(len(str(headers[ci-1])), max((len(str(r[ci-1])) for r in rows[:50] if ci-1 < len(r) and r[ci-1] is not None), default=0))
        ws.column_dimensions[get_column_letter(ci)].width = min(mx+4, 40)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    ws.freeze_panes = "A2"

def df_rows(df, cols): return [[row.get(c,"") for c in cols] for _, row in df.iterrows()]

# =============================================================================
# LOAD & PREPARE
# =============================================================================
def load_and_prepare():
    print("\n" + "="*60 + "\nLOADING & PREPARING DATA\n" + "="*60)
    cust = pd.read_excel(CUSTOMER_FILE, dtype=str)
    rm = pd.read_excel(RM_MASTER_FILE, dtype=str)
    t_data = pd.read_excel(T_DATA_FILE, dtype=str)
    bl = pd.read_excel(BRANCH_LOOKUP_FILE, dtype=str)
    for n, d in [("Customer",cust),("RM",rm),("T_Data",t_data),("Branch Lookup",bl)]:
        print(f"  {n:15}: {len(d):,} rows")

    # Clean strings only — no numeric conversion
    for c in [C["ck"],C["cf"],C["cc"],C["ef"],C["bc"],C["po"],C["rk"],C["cs"],C["qb"]]:
        if c in cust.columns: cust[c] = clean_str(cust[c])
    for c in [C["oc"],C["rc"],C["rb"],C["cc2"],C["tf"]]:
        if c in rm.columns: rm[c] = clean_str(rm[c])
    t_data[C["ck"]] = clean_str(t_data[C["ck"]])
    t_data[C["tb"]] = clean_str(t_data[C["tb"]])
    if C["tbf"] in t_data.columns:
        t_data[C["tbf"]] = clean_str(t_data[C["tbf"]])
        t_data[C["tb"]] = t_data[C["tb"]].replace("",np.nan).fillna(t_data[C["tbf"]]).fillna("")
    for c in [C["bcc"],C["bbc"],C["bso"]]:
        bl[c] = clean_str(bl[c])

    # Build branch lookup maps — keys are raw strings as they appear in the data
    cc2br = dict(zip(bl[C["bcc"]], bl[C["bbc"]]))
    sec2br = dict(zip(bl[C["bso"]], bl[C["bbc"]]))
    valid_br = set(bl[C["bbc"]].unique()) - {""}

    # RM true branch resolution: cost center -> branch, fallbacks
    rm["RM_Branch"] = rm[C["cc2"]].map(cc2br).fillna("")
    no = rm["RM_Branch"]==""
    rm.loc[no,"RM_Branch"] = rm.loc[no, C["cc2"]].where(rm.loc[no, C["cc2"]].isin(valid_br), "")
    no2 = rm["RM_Branch"]==""
    rm.loc[no2,"RM_Branch"] = rm.loc[no2, C["cc2"]].map(sec2br).fillna("")
    # Fallback: try BranchCode directly
    no3 = rm["RM_Branch"]==""
    rm.loc[no3,"RM_Branch"] = rm.loc[no3, C["rb"]].where(rm.loc[no3, C["rb"]].isin(valid_br), "")
    print(f"  RM branch resolved: {(rm['RM_Branch']!='').sum()} of {len(rm)}")

    # Customer true branch from T_Data
    def resolve_br(val):
        val = str(val).strip()
        if not val: return ""
        if val in cc2br: return cc2br[val]
        if val in valid_br: return val
        if val in sec2br: return sec2br[val]
        return ""

    tb_map = dict(zip(t_data[C["ck"]], t_data[C["tb"]]))
    cust["True_Branch"] = cust[C["ck"]].map(tb_map).fillna("").apply(resolve_br)

    # Balances
    for c in [C["bd"],C["bl"],C["be"],C["bi"]]: cust[c] = clean_num(cust[c])
    cust["TRV"] = cust[[C["bd"],C["bl"],C["be"],C["bi"]]].sum(axis=1)
    cust["SB"] = cust[C["bd"]] + cust[C["be"]] + cust[C["bi"]]
    cust["XS"] = clean_num(cust[C["cs"]]).astype(int)
    cust["Rel_Role"] = np.where(cust[C["rk"]]==cust[C["ck"]], "Anchor", "Member")

    # RM lookups
    rmd = rm.drop_duplicates(subset=[C["oc"]], keep="first")
    rm_role = dict(zip(rmd[C["oc"]], rmd[C["rc"]]))
    rm_br = dict(zip(rmd[C["oc"]], rmd["RM_Branch"]))
    rm_term = dict(zip(rmd[C["oc"]], rmd[C["tf"]].str.upper()))
    cust["RM_RC"] = cust[C["po"]].map(rm_role).fillna("")
    cust["RM_RT"] = cust["RM_RC"].apply(role_tier)
    cust["RM_BR"] = cust[C["po"]].map(rm_br).fillna("")
    cust["RM_TM"] = cust[C["po"]].map(rm_term).fillna("")

    # Exclusions — compare raw strings, no normalization
    ofc = cust[C["po"]]
    cc_up = cust[C["cc"]].str.upper()
    ef_up = cust[C["ef"]].str.upper()
    qb = cust[C["qb"]]
    excl = pd.DataFrame(index=cust.index)
    excl["qb"] = np.where(qb!="", "Has QB Officer", "")
    excl["cl"] = np.where(cc_up.isin(["PERSONAL","COMMERCIAL"]), "", "Non-Eligible Class: "+cc_up)
    excl["em"] = np.where(ef_up.isin(["Y","YES","TRUE","1"]), "Employee", "")
    excl["ct"] = np.where(ofc==CONTROL_CODE, "Control Code", "")
    excl["ec"] = np.where(ofc==EMPLOYEE_CODE, "Employee Code", "")
    cust["Excl_Reason"] = excl.apply(lambda r: "; ".join(v for v in r if v), axis=1)
    cust["Is_Excl"] = cust["Excl_Reason"]!=""

    elig = cust[~cust["Is_Excl"]].copy()
    elig["Need_UA"] = elig[C["cf"]].fillna("").astype(str).str.strip().str.upper()!="Y"

    # Tiers
    pers = elig[elig[C["cc"]].fillna("").astype(str).str.strip().str.upper()=="PERSONAL"]
    pa = pers[pers[C["cf"]].fillna("").astype(str).str.strip().str.upper()=="Y"]
    pw = pa[pa["TRV"]>0]
    bh = pw[pw["TRV"]<HNW_THRESHOLD]["TRV"].tolist()
    tiers = {}
    if len(bh)>=5:
        brks = jenkspy.jenks_breaks(bh, n_classes=5)
        tn = [("Low","SSA"),("Low-Mid","SSA"),("Mid","SSB"),("Upper-Mid","SSB"),("High","SSB")]
        td = [("Zero Balance",0,0)]
        for i,(n,_) in enumerate(tn):
            td.append((n, brks[i], brks[i+1] if i<4 else HNW_THRESHOLD))
        td.append(("HNW", HNW_THRESHOLD, float("inf")))
        tiers["P"] = td
        print(f"\n  Jenks breaks: {[round(b,2) for b in brks]}")
    else:
        tiers["P"] = [("Zero Balance",0,0),("Low",0.01,10000),("Low-Mid",10000,50000),
                      ("Mid",50000,100000),("Upper-Mid",100000,250000),("High",250000,HNW_THRESHOLD),
                      ("HNW",HNW_THRESHOLD,float("inf"))]

    def get_tier(row):
        cc_val = str(row[C["cc"]]).strip().upper()
        tv = row["TRV"]
        if cc_val=="COMMERCIAL": return "Zero Balance" if tv==0 else "All Commercial"
        if tv==0: return "Zero Balance"
        if tv>=HNW_THRESHOLD: return "HNW"
        for n, lo, hi in tiers.get("P",[])[1:-1]:
            if lo<=tv<hi: return n
        return "Unclassified"

    def get_target(row):
        cc_val = str(row[C["cc"]]).strip().upper()
        tier, xs = row["Tier"], row["XS"]
        if cc_val=="COMMERCIAL":
            return "BM > BS > SSB"
        if tier=="Zero Balance": return "SSB" if xs>=3 else "SSA"
        if tier=="HNW":
            if str(row["RM_RT"]).strip()=="SSB" and str(row[C["po"]]).strip() not in ALL_SPECIAL_CODES:
                return "SSB"
            return "BM > BS"
        if tier in ["Low","Low-Mid"]: return "SSA"
        if tier in ["Mid","Upper-Mid","High"]: return "SSB"
        return "SSA"

    elig["Tier"] = elig.apply(get_tier, axis=1)
    elig["Target"] = elig.apply(get_target, axis=1)

    print(f"  Total: {len(cust):,}  Excluded: {cust['Is_Excl'].sum():,}  Eligible: {len(elig):,}")
    for lbl, s in [("Tier",elig["Tier"]),("Target",elig["Target"])]:
        print(f"\n  {lbl}:")
        for v, n in s.value_counts().items(): print(f"    {v}: {n:,}")

    return cust, elig, rm, bl, valid_br

# =============================================================================
# BUILD CAPACITY
# =============================================================================
def build_capacity(rm, valid_br, bl):
    print("\n" + "="*60 + "\nBUILDING CAPACITY MODEL\n" + "="*60)
    vacancy_ocs = list(VACANCY_CODES.keys())

    # Regular officers: deduplicate by officer code, use cost-center-resolved branch
    rm_regular = rm[~rm[C["oc"]].isin(vacancy_ocs)].drop_duplicates(subset=[C["oc"]], keep="first").copy()

    # Vacancy officers: keep all rows, use BranchCode directly, deduplicate per branch
    rm_vacant = rm[rm[C["oc"]].isin(vacancy_ocs)].copy()
    rm_vacant["RM_Branch"] = rm_vacant[C["rb"]].where(rm_vacant[C["rb"]].isin(valid_br), "")
    rm_vacant["_uid"] = rm_vacant[C["oc"]] + "_" + rm_vacant["RM_Branch"]
    rm_vacant = rm_vacant.drop_duplicates(subset=["_uid"], keep="first")

    rmd = pd.concat([rm_regular, rm_vacant], ignore_index=True)
    rmd["RT"] = rmd[C["rc"]].apply(role_tier)
    term = rmd[C["tf"]].fillna("").str.upper()
    br = rmd["RM_Branch"].fillna("").astype(str).str.strip()

    staff = rmd[rmd[C["rc"]].isin(STAFFED_ROLES) & (~term.isin(["Y","YES","TRUE","1"])) &
               (~rmd[C["oc"]].isin(ALL_SPECIAL_CODES)) & (br!="") & (br.isin(valid_br))].copy()

    cap = {}
    for _, r in staff.iterrows():
        b, rt = str(r["RM_Branch"]).strip(), r["RT"]
        oc = str(r[C["oc"]]).strip()
        # Unique key for vacancy officers to avoid collisions across branches
        oc_key = f"{oc}_{b}" if oc in vacancy_ocs else oc
        cap.setdefault(b,{}).setdefault(rt,[]).append({
            "oc":oc_key, "on":str(r.get(C["on"],"")).strip(),
            "rc":str(r[C["rc"]]).strip(), "ac":0,
            "display_oc":oc})  # original code for output display

    tot = sum(len(o) for b in cap for o in cap[b].values())
    tc = sum(CAP.get(rt,0)*len(ol) for b in cap for rt,ol in cap[b].items())
    print(f"  Branches: {len(cap):,}  Officers: {tot:,}  Total capacity: {tc:,}")
    return cap

# =============================================================================
# ASSIGNMENT ENGINE
# =============================================================================
def find_rm(cap, br, primary, fallbacks=None):
    for role in [primary] + (fallbacks or []):
        officers = cap.get(br,{}).get(role,[])
        if not officers: continue
        limit = CAP.get(role, 999999)
        flag = "" if role==primary else f"Cascaded to {role}"
        under = [o for o in officers if o["ac"]<limit]
        if under:
            best = min(under, key=lambda x: x["ac"])
            return best["oc"], best.get("display_oc",best["oc"]), best["on"], role, best["rc"], flag
        continue
    return None, None, None, None, None, f"No capacity at branch {br}"

def incr(cap, br, role, oc):
    for o in cap.get(br,{}).get(role,[]):
        if o["oc"]==oc: o["ac"]+=1; return

# =============================================================================
# PROCESS ASSIGNMENTS
# =============================================================================
def process(elig, cap, valid_br):
    print("\n" + "="*60 + "\nPROCESSING ASSIGNMENTS\n" + "="*60)

    # Officer lookup from capacity
    olookup = {}
    for br in cap:
        for rt in cap[br]:
            for o in cap[br][rt]:
                olookup[o["oc"]] = {"br":br,"rt":rt,"rc":o["rc"],"on":o["on"],"display_oc":o.get("display_oc",o["oc"])}

    all_c = elig.sort_values("SB", ascending=False)
    recs, pending = [], []

    # Pass 1A: Validate existing
    print("  Pass 1A: Validating existing assignments...")
    for _, row in all_c.iterrows():
        ofc = str(row.get(C["po"],"")).strip()
        cb = str(row.get(C["bc"],"")).strip()
        tb = str(row.get("True_Branch","")).strip()
        target = str(row.get("Target","")).strip()
        tm = str(row.get("RM_TM","")).strip()

        # Rule 7: target branch — validate current branch against branch lookup
        if cb and cb in valid_br:
            target_br = cb
        elif tb and tb in valid_br:
            target_br = tb
        else:
            target_br = ""

        rec = {"CK":str(row.get(C["ck"],"")).strip(), "CN":str(row.get(C["cn"],"")).strip(),
               "CC":str(row.get(C["cc"],"")).strip(), "RK":str(row.get(C["rk"],"")).strip(),
               "RR":str(row.get("Rel_Role","")).strip(), "Tier":str(row.get("Tier","")).strip(),
               "TRV":row.get("TRV",0), "SB":row.get("SB",0),
               "Dep":row.get(C["bd"],0), "Loan":row.get(C["bl"],0),
               "Eq":row.get(C["be"],0), "Ins":row.get(C["bi"],0),
               "Cur_Ofc":ofc, "Cur_Br":cb, "True_Br":tb, "Tgt_Br":target_br,
               "Rec_Ofc":None,"Rec_Ofc_Display":None,"Rec_Name":None,"Rec_RC":None,"Rec_Br":target_br,
               "Move":"Pending","Reason":"","Flag":"","Target":target}

        # No valid branch available
        if not target_br:
            flag_msg = "Current branch not in branch lookup" if cb else "No current branch"
            if not tb:
                flag_msg += " and no true branch"
            elif tb not in valid_br:
                flag_msg += f" and true branch {tb} not in branch lookup"
            rec.update(Rec_Ofc="REVIEW", Move="Needs Review", Reason="No valid branch available",
                       Flag=flag_msg)
            rec["Rec_Ofc_Display"] = "REVIEW"
            recs.append(rec); continue

        # Flag if using true branch instead of current
        if cb and cb not in valid_br:
            rec["Flag"] = f"Current branch {cb} not in branch lookup; using true branch {target_br}"
        elif not cb:
            rec["Flag"] = "No current branch; assigned via true branch"

        # Rule 3: Cust Flag null
        if row.get("Need_UA", False):
            rec.update(Rec_Ofc="Unassigned", Move="Unassignment", Reason="Cust Flag not Y")
            rec["Rec_Ofc_Display"] = "Unassigned"
            recs.append(rec); continue

        # Check if current assignment valid
        valid = False
        if ofc and ofc not in ALL_SPECIAL_CODES and tm!="Y" and ofc in olookup:
            info = olookup[ofc]
            expected = EXPECTED_TIERS.get(target,[])
            if info["rt"] in expected and info["br"]==target_br:
                valid = True

        if valid:
            incr(cap, info["br"], info["rt"], ofc)
            rec.update(Rec_Ofc=ofc, Rec_Name=info["on"], Rec_RC=info["rc"],
                       Move="No Change", Reason="Already correctly assigned")
            rec["Rec_Ofc_Display"] = info.get("display_oc", ofc)
            recs.append(rec)
        else:
            recs.append(rec)
            pending.append(len(recs)-1)

    kept = sum(1 for r in recs if r["Move"]=="No Change")
    print(f"  Kept: {kept:,}  Need assignment: {len(pending):,}")

    # Pass 1B: Assign pending by highest balance
    print("  Pass 1B: Assigning to open slots...")
    pending.sort(key=lambda i: recs[i]["SB"], reverse=True)
    asgn, unasgn = 0, 0

    for i in pending:
        rec = recs[i]
        target_br = rec["Tgt_Br"]
        target = rec["Target"]
        primary, fb = ROLE_CASCADE.get(target, ("SSA",[]))

        oc_key, display_oc, on, rt, rc, flag = find_rm(cap, target_br, primary, fb)
        if oc_key:
            incr(cap, target_br, rt, oc_key)
            rec["Rec_Ofc"], rec["Rec_Name"], rec["Rec_RC"] = oc_key, on, rc
            rec["Rec_Ofc_Display"] = display_oc
            if display_oc==rec["Cur_Ofc"] and target_br==rec["Cur_Br"]:
                rec["Move"], rec["Reason"] = "No Change", "Already correctly assigned"
            elif target_br!=rec["Cur_Br"]:
                rec["Move"] = "Branch + RM Change"
                rec["Reason"] = f"Branch {target_br}, was {rec['Cur_Br']}"
            else:
                rec["Move"], rec["Reason"] = "RM Change", f"Reassigned to {rc} per {rec['Tier']}"
            if flag: rec["Flag"] = (rec["Flag"]+"; "+flag) if rec["Flag"] else flag
            asgn += 1
        else:
            rec.update(Rec_Ofc="Unassigned", Move="Unassignment", Reason=f"No capacity at {target_br}")
            rec["Rec_Ofc_Display"] = "Unassigned"
            if flag: rec["Flag"] = (rec["Flag"]+"; "+flag) if rec["Flag"] else flag
            unasgn += 1

    print(f"  Assigned: {asgn:,}  Unassigned: {unasgn:,}")
    rdf = pd.DataFrame(recs)

    # Build officer info for Pass 2/3
    oinfo = {}
    for br in cap:
        for rt in cap[br]:
            lim = CAP.get(rt, 999999)
            for o in cap[br][rt]:
                oinfo[o["oc"]] = {"br":br,"rt":rt,"rc":o["rc"],"cap":lim,"on":o["on"],
                                  "display_oc":o.get("display_oc",o["oc"])}

    # Pass 2: Trim & Fill
    print("\n  Pass 2: Trim over-capacity, fill under-capacity...")
    trimmed = 0
    for oc, info in oinfo.items():
        mask = rdf["Rec_Ofc"]==oc
        excess = mask.sum() - info["cap"]
        if excess<=0: continue
        lowest = rdf.loc[mask].sort_values("SB").head(excess).index
        rdf.loc[lowest, ["Rec_Ofc","Rec_Name","Rec_RC"]] = ["Unassigned","",""]
        rdf.loc[lowest, "Rec_Ofc_Display"] = "Unassigned"
        rdf.loc[lowest, "Move"] = "Unassignment"
        rdf.loc[lowest, "Reason"] = "Trimmed - lowest balance in over-capacity portfolio"
        rdf.loc[lowest, "Flag"] = rdf.loc[lowest,"Flag"].apply(lambda f: (f+"; Pass 2 trim") if f else "Pass 2 trim")
        trimmed += excess
    print(f"  Trimmed: {trimmed:,}")

    filled = 0
    for oc, info in oinfo.items():
        slots = info["cap"] - (rdf["Rec_Ofc"]==oc).sum()
        if slots<=0: continue
        allowed = [t for t, tiers in EXPECTED_TIERS.items() if info["rt"] in tiers]
        ua = rdf[(rdf["Rec_Ofc"]=="Unassigned")&(rdf["Rec_Br"]==info["br"])&(rdf["Target"].isin(allowed))].sort_values("SB",ascending=False)
        fill_idx = ua.head(slots).index
        if len(fill_idx)==0: continue
        rdf.loc[fill_idx, ["Rec_Ofc","Rec_Name","Rec_RC"]] = [oc, info["on"], info["rc"]]
        rdf.loc[fill_idx, "Rec_Ofc_Display"] = info["display_oc"]
        rdf.loc[fill_idx, "Move"] = "RM Change"
        rdf.loc[fill_idx, "Reason"] = "Filled under-capacity (highest balance)"
        rdf.loc[fill_idx, "Flag"] = rdf.loc[fill_idx,"Flag"].apply(lambda f: (f+"; Pass 2 fill") if f else "Pass 2 fill")
        filled += len(fill_idx)
    print(f"  Filled: {filled:,}")

    # Pass 3: Swap
    print("\n  Pass 3: Swap low-balance assigned for high-balance unassigned...")
    swapped = 0
    go = True
    while go:
        go = False
        touched = set()
        for oc, info in oinfo.items():
            amask = (rdf["Rec_Ofc"]==oc)&(~rdf.index.isin(touched))
            if amask.sum()==0: continue
            low = rdf.loc[amask].sort_values("SB").iloc[0]
            allowed = [t for t, tiers in EXPECTED_TIERS.items() if info["rt"] in tiers]
            umask = (rdf["Rec_Ofc"]=="Unassigned")&(rdf["Rec_Br"]==info["br"])&(~rdf.index.isin(touched))&(rdf["Target"].isin(allowed))
            if umask.sum()==0: continue
            high = rdf.loc[umask].sort_values("SB",ascending=False).iloc[0]
            if high["SB"]>low["SB"]:
                li, hi = low.name, high.name
                rdf.loc[li, ["Rec_Ofc","Rec_Name","Rec_RC"]] = ["Unassigned","",""]
                rdf.loc[li, "Rec_Ofc_Display"] = "Unassigned"
                rdf.loc[li, "Move"], rdf.loc[li,"Reason"] = "Unassignment", "Swapped out for higher-balance customer"
                rdf.loc[li, "Flag"] = (rdf.loc[li,"Flag"]+"; Pass 3 swap out") if rdf.loc[li,"Flag"] else "Pass 3 swap out"
                rdf.loc[hi, ["Rec_Ofc","Rec_Name","Rec_RC"]] = [oc, info["on"], info["rc"]]
                rdf.loc[hi, "Rec_Ofc_Display"] = info["display_oc"]
                rdf.loc[hi, "Move"], rdf.loc[hi,"Reason"] = "RM Change", "Swapped in (higher balance)"
                rdf.loc[hi, "Flag"] = (rdf.loc[hi,"Flag"]+"; Pass 3 swap in") if rdf.loc[hi,"Flag"] else "Pass 3 swap in"
                touched.update([li, hi])
                swapped += 1; go = True
    print(f"  Swapped: {swapped:,}")

    print(f"\n  Final:")
    for t, n in rdf["Move"].value_counts().items(): print(f"    {t}: {n:,}")
    return rdf

# =============================================================================
# WRITE OUTPUT
# =============================================================================
def write_output(rdf, cust, cap):
    print("\n" + "="*60 + "\nWRITING OUTPUT\n" + "="*60)
    wb = Workbook()

    # Sheet 1: Recommendations
    ws = wb.active; ws.title = "Recommendations"
    rc = ["CK","CN","CC","RK","RR","Tier","TRV","SB","Loan","Dep","Eq","Ins",
          "Cur_Ofc","Cur_Br","True_Br","Rec_Ofc_Display","Rec_Name","Rec_RC","Rec_Br","Move","Reason","Flag"]
    rh = ["Cust Key","Cust Name","Cust Class","Relationship Key","Relationship Role",
          "Balance Tier","Total Relationship Value","Assets","Liabilities","Deposits","Equity","Insurance",
          "Current Officer","Current Branch","True Branch","Recommended Officer","Recommended Officer Name",
          "Recommended Role","Recommended Branch","Movement Type","Reason","Flag"]
    write_sheet(ws, rh, df_rows(rdf, rc))

    # Sheet 2: Movement Summary
    ms = rdf["Move"].value_counts().reset_index(); ms.columns=["Movement Type","Count"]
    write_sheet(wb.create_sheet("Movement Summary"), ["Movement Type","Count"], ms.values.tolist(), FILLS["green"])

    # Sheet 3: Branch Impact
    bb = rdf["Cur_Br"].value_counts().rename("Before")
    ba = rdf["Rec_Br"].value_counts().rename("After")
    bi = pd.concat([bb,ba],axis=1).fillna(0).astype(int)
    bi["Net"] = bi["After"]-bi["Before"]
    bi = bi.reset_index().rename(columns={"index":"Branch"}).sort_values("Branch")
    write_sheet(wb.create_sheet("Branch Impact"), ["Branch","Before","After","Net Change"], bi.values.tolist())

    # Sheet 4: RM Impact
    rb = rdf["Cur_Ofc"].value_counts().rename("Before")
    ro = rdf["Rec_Ofc"].fillna("").astype(str)
    va = rdf[~ro.isin(["REVIEW","Unassigned","None",""])]
    ra = va["Rec_Ofc"].value_counts().rename("After")
    ri = pd.concat([rb,ra],axis=1).fillna(0).astype(int)
    ri["Net"] = ri["After"]-ri["Before"]
    ri = ri.reset_index().rename(columns={"index":"Officer"}).sort_values("Officer")
    write_sheet(wb.create_sheet("RM Impact"), ["Officer","Before","After","Net Change"], ri.values.tolist())

    # Sheet 5: Exceptions
    exc = rdf[rdf["Move"]=="Needs Review"]
    ec = ["CK","CN","Tier","True_Br","Cur_Ofc","Cur_Br","Reason","Flag"]
    eh = ["Cust Key","Cust Name","Balance Tier","True Branch","Current Officer","Current Branch","Reason","Flag"]
    write_sheet(wb.create_sheet("Exceptions"), eh, df_rows(exc, ec), FILLS["red"])

    # Sheet 6: Exclusion Summary
    exd = cust[cust["Is_Excl"]]
    xc = [C["ck"],C["cn"],C["cc"],C["bc"],C["po"],"Excl_Reason"]
    xh = ["Cust Key","Cust Name","Cust Class","Branch Code","Primary Officer","Exclusion Reason"]
    write_sheet(wb.create_sheet("Exclusion Summary"), xh, df_rows(exd, xc), FILLS["orange"])

    # Sheet 7: Rebalancing Candidates
    ro7 = rdf["Rec_Ofc"].fillna("").astype(str)
    ac = rdf[~ro7.isin(["REVIEW","Unassigned","None",""])].sort_values("SB")
    rc7 = ["CK","CN","CC","Tier","SB","Loan","TRV","Dep","Eq","Ins","Cur_Br","Cur_Ofc",
           "Rec_Ofc_Display","Rec_Name","Rec_RC","Rec_Br","Reason","Flag"]
    rh7 = ["Cust Key","Cust Name","Cust Class","Balance Tier","Assets","Liabilities",
           "Total Relationship Value","Deposits","Equity","Insurance","Current Branch","Current Officer",
           "Recommended Officer","Recommended Officer Name","Recommended Role","Recommended Branch","Reason","Flag"]
    write_sheet(wb.create_sheet("Rebalancing Candidates"), rh7, df_rows(ac, rc7), FILLS["green"])

    # Sheet 8: Branch Role Summary
    brs = []
    for br in sorted(cap):
        for rt in sorted(cap[br]):
            lim = CAP.get(rt,0)
            ocs = [o["oc"] for o in cap[br][rt]]
            n = len(ocs); tc = lim*n
            mask = rdf["Rec_Ofc"].isin(ocs)
            cnt = int(mask.sum())
            assets = float(rdf.loc[mask,"SB"].sum())
            loans = float(rdf.loc[mask,"Loan"].sum())
            brs.append({"Br":br,"Role":rt,"Num":n,"CapPer":lim,"TotCap":tc,
                        "Cnt":cnt,"Open":tc-cnt,"Assets":round(assets,2),"Liabilities":round(loans,2)})
    bdf = pd.DataFrame(brs)
    write_sheet(wb.create_sheet("Branch Role Summary"),
                ["Branch","Role","Num Officers","Cap Per Officer","Total Capacity","Customer Count","Open Slots","Assets","Liabilities"],
                df_rows(bdf, ["Br","Role","Num","CapPer","TotCap","Cnt","Open","Assets","Liabilities"]))

    # Sheet 9: RM Portfolio Summary
    rps = []
    for br in sorted(cap):
        for rt in sorted(cap[br]):
            lim = CAP.get(rt,0)
            for o in cap[br][rt]:
                m = rdf["Rec_Ofc"]==o["oc"]
                cnt = int(m.sum())
                assets = float(rdf.loc[m,"SB"].sum())
                loans = float(rdf.loc[m,"Loan"].sum())
                rps.append({"Br":br,"OC":o.get("display_oc",o["oc"]),"ON":o["on"],"Role":rt,"RC":o["rc"],
                            "Cap":lim,"Cnt":cnt,"Open":lim-cnt,"Assets":round(assets,2),"Liabilities":round(loans,2)})
    rpdf = pd.DataFrame(rps)
    write_sheet(wb.create_sheet("RM Portfolio Summary"),
                ["Branch","Officer Code","Officer Name","Role","Role Code","Capacity","Customer Count","Open Slots","Assets","Liabilities"],
                df_rows(rpdf, ["Br","OC","ON","Role","RC","Cap","Cnt","Open","Assets","Liabilities"]))

    for n, c in [("Recommendations",len(rdf)),("Exceptions",len(exc)),("Exclusions",len(exd)),
                 ("Rebalancing",len(ac)),("Branch Roles",len(bdf)),("RM Portfolios",len(rpdf))]:
        print(f"  {n}: {c:,}")

    wb.save(OUTPUT_FILE)
    print(f"\n  Saved to {OUTPUT_FILE}")

# =============================================================================
# MAIN
# =============================================================================
def main():
    print("\n" + "="*60 + "\nPORTFOLIO REBALANCING ALGORITHM\n" + "="*60)
    cust, elig, rm, bl, valid_br = load_and_prepare()
    cap = build_capacity(rm, valid_br, bl)
    rdf = process(elig, cap, valid_br)
    write_output(rdf, cust, cap)
    print("\n" + "="*60 + "\nCOMPLETE\n" + "="*60)

if __name__ == "__main__":
    main()
